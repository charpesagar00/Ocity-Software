from customtkinter import *
import tkinter.messagebox as msg 
import os , sys
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font , Alignment
import smtplib , shutil 
import webbrowser
from PIL import Image
import random , logging , requests
home = os.path.expanduser("~")
os.chdir(home) # ---------> C:\Users\Administrator
if not os.path.exists("KCA-Student Login"):
    os.mkdir("KCA-Student Login")

# Current Date and Time Function
current_date = ""
current_time = ""
def get_date_time():
    global current_time, current_date, ist_time
    try:
        from email.utils import parsedate_to_datetime
        from datetime import datetime , timedelta
        response = requests.get("https://google.com",timeout=5)
        gmt_time = parsedate_to_datetime(response.headers.get("date"))
        ist_time = (gmt_time + timedelta(hours=5,minutes=30)).replace(tzinfo=None)
        current_time = ist_time.strftime("%I:%M:%S %p")
        current_date = ist_time.strftime("%Y-%m-%d")
    except:
        current_date = "-"
        current_time = "-"

# ----------------------------------------------------------------------------------------------
from email.utils import parsedate_to_datetime
from datetime import timedelta , datetime

def get_expiry_date():
    global expiry_date
    get_date_time()
    expiry_date = ist_time + timedelta(days=30,hours=0,minutes=0) # <---------- SET EXPIRY DATE HERE
    return expiry_date

# ---------------------------------------------------------------------------------------------
# To Get Listed on GitHub Server
import socket , getpass, uuid , json
from github import Github
TOKEN = "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"
REPO_NAME = "charpesagar00/Ocity-Software"

rec_email1 = None
get_date_time()
pc = {
    "Device Name" : socket.gethostname(),
    "User Name" : getpass.getuser(),
    "MAC ID" : uuid.getnode(),
    "Date Joined" : current_date,
    "Time Joined" : current_time,
    "Expiry Date" : None ,
    "Last Payment Date" : None,
    "Registered E-mail" : None
}

json_str = json.dumps(pc,indent=4)
g = Github(TOKEN)
repo = g.get_repo(REPO_NAME)
filename = f"Devices/PC:{socket.gethostname()}.json"
try:
    repo.get_contents(filename)

except:
    repo.create_file(path=filename,message="PC Info",content=json_str)

# ----------------------------------------------------------------------------------------------
os.chdir(os.path.join(home,"KCA-Student Login")) # ---------> C:\Users\Administrator\KCA-Student Login

# QR Code Window Code >>>
def relogin_after_expiry():
    def get_pay_email():
        global OTP , myotp , rec_email1
        email1 = "charpesagar00@gmail.com"
        rec_email1 = get.get()

        sym = [
        'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 
        'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 
        '0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9',
        '!', '#', '$', '%', '&', '*', '@', '^', '_' ]

        if rec_email1.endswith("@gmail.com"):
            OTP = ""
            for i in range(8):
                num = random.choice(sym)
                OTP = OTP + str(num)
            
            server = smtplib.SMTP("SMTP.gmail.com",587)
            server.starttls()

            m = f'''SUBJECT : Your OTP\n\n
            Your One Time Password (OTP) is : {OTP}
            Enjoy!
            User : {rec_email1}
            '''

            server.login(email1,"jqpe uqif xbzv fpqj")
            server.sendmail(email1,email1,m)
            server.close()
            x.grid(row=2,column=0,padx=(35,0),sticky="w")
            get_otp.grid(row=2,column=1,sticky="w",pady=5)
            
            y.destroy()
            z.grid(row=3,column=0,columnspan=2,pady=15)
            
        elif rec_email1 == "":
            msg.showwarning("Error","Please insert your valid email address.")

        else:
            msg.showwarning("Error","Please recheck your email.")

    def get_pay_otp():
            myotp = str(get_otp.get())
            if OTP == myotp :
                
                with open("payment_authentication.txt","w") as file:
                    file.write(f"Payment Completion Date : {current_date}")
                    file.write(f"Registered Email : {rec_email1}")

                contents = repo.get_contents(filename)
                data = json.loads(contents.decoded_content.decode("utf-8"))

                data["Expiry Date"] = str(get_expiry_date())
                data["Last Payment Date"] = str(current_date)
                data["Registered E-mail"] = rec_email1

                repo.update_file(path=filename,message="PC Info",content=json.dumps(data,indent=4),sha=contents.sha)
                msg.showinfo("Success!","Congratulations! OTP verificationn is Successfull!")
                my.destroy()

            else:
                if msg.askretrycancel("Error!","Invalid OTP. Please Enter Valid OTP."):
                    return
                else:
                    sys.exit()

    my = CTk()
    my.config(bg="#333333")
    my.title("Login First")
    my.geometry("420x460")

    qr = Image.open(os.path.join(home,"Desktop","pay_qr.jpeg"))
    my_image = CTkImage(light_image=qr,
                        dark_image=qr,
                        size=(300, 290)) 
    CTkLabel(my, image=my_image, text="").grid(row=0,column=0,columnspan=2,padx=60,pady=20,sticky="we")
    CTkLabel(my,text="Email",text_color="#ffffff",bg_color="#333333",font=("Calibri", 18)).grid(row=1,column=0,padx=(35,0),sticky="w")
    get = CTkEntry(my,font=("Calibri", 18),fg_color="#333333",bg_color="#333333",text_color="#ffffff",width=280)
    get.grid(row=1,column=1,sticky="w")

    x = CTkLabel(my,text="OTP",text_color="#ffffff",bg_color="#333333",font=("Calibri", 18))
    get_otp = CTkEntry(my,font=("Calibri", 18),fg_color="#333333",bg_color="#333333",text_color="#ffffff",width=200)
   
    y = CTkButton(my,text="Get OTP",font=("Calibri", 20,"bold"),command=get_pay_email,bg_color="#333333",corner_radius=5)
    y.grid(row=3,column=0,columnspan=2,pady=15)

    z = CTkButton(my,text="Submit",font=("Calibri", 20,"bold"),command=get_pay_otp,bg_color="#333333",corner_radius=5)

    def on_closing1():
        sys.exit()
    my.protocol("WM_DELETE_WINDOW",on_closing1)
    my.mainloop()
# -----------------------------------------------------------------------------------------------------------------------------------------------------------------------

if not os.path.exists("payment_authentication.txt"):
    relogin_after_expiry() # Also Used for first time and after for expiry date

try :
    contents = repo.get_contents(filename)
    data = json.loads(contents.decoded_content.decode("utf-8"))
    expiry_date_str = data["Expiry Date"]
except:
    pass

get_date_time()
expiry_obj = datetime.strptime(expiry_date_str,"%Y-%m-%d %H:%M:%S")
current_obj = datetime.strptime(current_date,"%Y-%m-%d")

if current_obj > expiry_obj :
    if msg.askyesno("Expired!","Oops! Your subscription has expired.\nPlease renew your plan to continue using all features.\n\nClick yes if you want to continue."):
        relogin_after_expiry()
    else:
        sys.exit()

# -----------------------------------------------------------------------------------------------------------------------------------------------------
    # User Login Form
    # Login Window 
def login_form_window():
    def login_info():
        global email
        global passkey
        email = em.get()
        passkey = pk.get()

        if not email.endswith("@gmail.com"):
            msg.showwarning("Oops!","Maybe your Email is not valid")
            return

        # Open Excel File With Headings 
        if not os.path.exists("Student Enquiries Backup.xlsx"):
            wb = Workbook()
            headings1 = ["Sr.No","Name","Phone No","Gender","Academic Status","Class/Year","Board","Course","Total Offered Fees","DOB","College Name","Coaching Name","Parent's No","E-mail","Address","Date","Time"]
            s = wb.active
            s.title = "Admissions"
            s.append(headings1)
            s.merge_cells("r1:y1")
            s["r1"]="Installments"
            s.column_dimensions['A'].width = 7
            s.column_dimensions['B'].width = 30
            s.column_dimensions['C'].width = 13
            s.column_dimensions['D'].width = 8
            s.column_dimensions['E'].width = 16
            s.column_dimensions['F'].width = 20
            s.column_dimensions['G'].width = 10
            s.column_dimensions['H'].width = 23
            s.column_dimensions['I'].width = 17
            s.column_dimensions['J'].width = 17
            s.column_dimensions['K'].width = 35
            s.column_dimensions['L'].width = 35
            s.column_dimensions['M'].width = 13
            s.column_dimensions['N'].width = 30
            s.column_dimensions['O'].width = 50
            s.column_dimensions['P'].width = 12
            s.column_dimensions['Q'].width = 12

            for cell in s[1]:
                 cell.font = Font(bold=True)

            headings2 = ["Sr. No","Gender","Name","Course","Source","Qualification","Mobile No.","College","Address","Date","Time"]
            sheet2 = wb.create_sheet(title="Student Enquiries")
            sheet2.append(headings2)

            sheet2.column_dimensions['A'].width = 7
            sheet2.column_dimensions['B'].width = 10
            sheet2.column_dimensions['C'].width = 18
            sheet2.column_dimensions['D'].width = 25
            sheet2.column_dimensions['E'].width = 15
            sheet2.column_dimensions['F'].width = 20
            sheet2.column_dimensions['G'].width = 15
            sheet2.column_dimensions['H'].width = 35
            sheet2.column_dimensions['I'].width = 40
            sheet2.column_dimensions['J'].width = 12
            sheet2.column_dimensions['K'].width = 12

            for cell in sheet2[1]:
                cell.font = Font(bold=True)
            wb.save("Student Enquiries Backup.xlsx")

        with open("login_creadentials.txt","w") as file :
            file.write(f"{email}|{passkey}\n")

        msg.showinfo("Email Updated!","Your Email Address is Updated Sucessfully!")
        start.destroy()

    # Starting Login Form
    start = CTk()
    start.title("Login Form")
    start.resizable(FALSE,FALSE)
    start.config(bg="#333333")
        
    login = CTkScrollableFrame(start,width=410,height=140,label_text="Login Creadentials",label_text_color="#FFFFFF",label_fg_color="#333333",label_font=("Montserrat",16),border_color="#97ECFF",border_width=2,fg_color="#ECFFFF",bg_color="#333333",corner_radius=12)
    login.grid(row=0,column=0,padx=10,pady=10)

    CTkLabel(login,text="Email",fg_color="#FFC5C5",text_color="#000000",bg_color="#ECFFFF",padx=19,corner_radius=18,font=("Montserrat",14)).grid(row=0,column=0,sticky="wn",padx=15,pady=(0,10))
    em = CTkEntry(login,width=250,corner_radius=8,fg_color="#333333",text_color="#ffffff",border_width=1,border_color="#FFC5C5")
    em.grid(row=0,column=1,pady=(0,10))

    CTkLabel(login,text="Passkey",fg_color="#FFC5C5",text_color="#000000",bg_color="#ECFFFF",padx=10,corner_radius=18,font=("Montserrat",14)).grid(row=1,column=0,sticky="wn",padx=15)
    pk = CTkEntry(login,width=250,corner_radius=8,fg_color="#333333",text_color="#ffffff",border_width=1,border_color="#FFC5C5")
    pk.grid(row=1,column=1)

    CTkButton(login,text="Submit",font=("Montserrat",14),width=5,corner_radius=15,command=login_info).grid(row=2,column=0,columnspan=2,pady=(20,10),sticky="ew")

    CTkLabel(login,text="Generate Passkey : ",text_color="#000000").grid(row=3,column=0,sticky="w",pady=(0,5))
    link = CTkLabel(login,text="https://myaccount.google.com/apppasswords",text_color="#0040B7",text_color_disabled="#80E1F2")
    link.grid(row=3,column=1,columnspan=2,sticky="w",pady=(0,5))
    link.bind("<Button-1>",lambda x : webbrowser.open_new_tab("https://myaccount.google.com/apppasswords"))
    
    CTkLabel(login,text='Steps ⬇ :  ',text_color="#000000",justify="left",font=("Calibri",18,"bold")).grid(row=4,column=0,columnspan=2,sticky="w",pady=5)
    CTkLabel(login,text='>>  Click on Link',text_color="#000000",justify="left").grid(row=5,column=0,columnspan=2,sticky="w",padx=35)
    CTkLabel(login,text='>>  Login your account',text_color="#000000",justify="left").grid(row=6,column=0,columnspan=2,sticky="w",padx=35)
    CTkLabel(login,text='>>  Type "KCA" in App name',text_color="#000000",justify="left").grid(row=7,column=0,columnspan=2,sticky="w",padx=35)
    CTkLabel(login,text='>>  Click "Create"',text_color="#000000",justify="left").grid(row=8,column=0,columnspan=2,sticky="w",padx=35)
    CTkLabel(login,text='>>  Copy the passkey and paste it in "Passkey" field',text_color="#000000",justify="left").grid(row=9,column=0,columnspan=2,sticky="w",padx=35)
    CTkLabel(login,text='>>  Click on Submit',text_color="#000000",justify="left").grid(row=10,column=0,columnspan=2,sticky="w",padx=35)

    def on_closing2():
           sys.exit()
    start.protocol("WM_DELETE_WINDOW",on_closing2)
    start.mainloop()

if not os.path.exists("login_creadentials.txt"):
    login_form_window() # Login Form Window Calls
# -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Version Checker System
if not os.path.exists("version.json"):
    with open("version.json", "w") as file:
        json.dump({"current_version": "0.0"}, file) # <<<------------------ SET VERSION HERE...

import urllib.request
with open("version.json", "r") as file:
    local_data = json.load(file)
    current_version = local_data.get("current_version")

try:
    headers = {"Authorization" : f"token {TOKEN}"}
    response = requests.get("https://raw.githubusercontent.com/charpesagar00/Ocity-Software/refs/heads/main/Latest%20Version.json?",headers=headers)
    data = response.json()
except requests.exceptions.RequestException as e:
    print(f"Network error checking for updates: {e}")
except Exception as e:
    print(f"An unexpected error occurred during update: {e}")

latest_version = data["version"]
download_url = data["download_url"]
print(latest_version)
print(download_url)

if latest_version > current_version:

    choice = msg.askyesno("Update Available!", f"A new version ({latest_version}) is available! Would you like to download it now?")
    if choice == True:
        if download_url != "":
            urllib.request.urlretrieve(download_url,os.path.join(home,"Downloads","QR.Code.exe"))
            msg.showinfo("Success", "Download completed! Please check your folder.")
            with open("version.json", "w") as file:
                json.dump({"current_version": latest_version}, file)
        else:
                print("Error: Download URL is missing from the remote JSON configuration.")
    else:
       print("Software is already up to date.")

# -------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Main Window Starts
root = CTk()
root.title("KAILASH COMPUTER ACADEMY ( Government Authorised Center )")
root.after(0, lambda: root.state('zoomed'))
root.config(bg="#333333")

start_frame = CTkFrame(root,fg_color="#333333",bg_color="#333333",border_width=0)
start_frame.grid(row=0,column=0,sticky="w")

def display_dashboard_frame():
    f1.grid_forget()
    f2.grid_forget()
    enq_frame.grid_forget()
    dashboard_frame.grid(row=1,column=0,pady=10,padx=10)

def display_admission_frame():
    dashboard_frame.grid_forget()
    enq_frame.grid_forget()
    f1.grid(row=1, column=0, sticky="ew", padx=10, pady=(5,5))
    f2.grid(row=2, column=0, sticky="new", padx=10, pady=0)

def display_enquiry_frame():
    dashboard_frame.grid_forget()
    f1.grid_forget()
    f2.grid_forget()
    enq_frame.grid(row=1,column=0,pady=(10,0),padx=10,sticky="w")

CTkButton(start_frame,text="Dashboard",font=("Calibri",20,"bold"),bg_color="#333333",fg_color="#bfffff",text_color="#333333",hover_color="#87FFFD",corner_radius=50,width=200,border_color="#39FFF5",border_width=1,command=display_dashboard_frame).grid(row=0,column=0,padx=10,pady=(10,5),sticky="w")
CTkButton(start_frame,text="Add Enquiry",font=("Calibri",20,"bold"),bg_color="#333333",fg_color="#bfffff",text_color="#333333",hover_color="#87FFFD",corner_radius=50,width=200,border_color="#39FFF5",border_width=1,command=display_enquiry_frame).grid(row=0,column=1,padx=10,pady=(10,5),sticky="w")
CTkButton(start_frame,text="Add Admission",font=("Calibri",20,"bold"),bg_color="#333333",fg_color="#bfffff",text_color="#333333",hover_color="#87FFFD",corner_radius=50,width=200,border_color="#39FFF5",border_width=1,command=display_admission_frame).grid(row=0,column=2,padx=10,pady=(10,5),sticky="w")

# Dashboard Frame
dashboard_frame = CTkScrollableFrame(root,height=500,bg_color="#333333",fg_color="#ecffff",width=1100,corner_radius=12,border_color="#21FFE1",border_width=1,label_text="Enquiries",label_font=("Montserrat",18))
dashboard_frame.grid(row=1,column=0,padx=10,pady=10)

# Filters to display Enquiries
def filter_frame_toggle():
    if tog.get():
        filter_enq_headings_frame.grid(row=1,column=0,columnspan=10,sticky="nsew", padx=5, pady=(5,30))
    else:
        filter_enq_headings_frame.grid_remove()


top_widgest_headings_frame = CTkFrame(dashboard_frame,width=1100,bg_color="#333333",fg_color="#ecffff",border_color="#00F6B1",border_width=1,corner_radius=12,background_corner_colors=("#ecffff","#ecffff","#ecffff","#ecffff"))
top_widgest_headings_frame.grid(row=0,column=0,columnspan=10,sticky="nsew", padx=5, pady=5)

filter_enq_headings_frame = CTkFrame(dashboard_frame,bg_color="#333333",fg_color="#ecffff",border_color="#333333",border_width=1.2,corner_radius=12,background_corner_colors=("#ecffff","#ecffff","#ecffff","#ecffff"))

tog = BooleanVar()
toggle = CTkCheckBox(top_widgest_headings_frame,text="Filter Column",text_color="#333333",font=("Calibri",14),fg_color="#00256A",bg_color="#ecffff",checkbox_width=20,checkbox_height=20,border_width=2,variable=tog,command=filter_frame_toggle)
toggle.grid(row=0,column=0,sticky="ns",padx=10,pady=10)
filter_enq_headings = [ "Sr.No","Gender","Name","Course","Source","Mobile No.","Date","Time"]

filters_on = []
active_filter_list = []
active_data_lables = []
selected_filters = []
button_lables = []

def edit_enquiry_function(temp):
    if msg.askyesno("Entry Deletation","Do you really want to delete this entry ? "):
        wb = load_workbook("Student Enquiries Backup.xlsx")
        sheet2 = wb["Student Enquiries"]
        sheet2.delete_rows(idx=temp+2,amount=1)
        wb.save("Student Enquiries Backup.xlsx")

        shutil.copy(os.path.join(home,"KCA-Student Login","Student Enquiries Backup.xlsx"),os.path.join(home,"Documents"))
        os.replace(os.path.join(home,"Documents","Student Enquiries.xlsx"),os.path.join(home,"Documents","Student Enquiries.xlsx"))
        msg.showinfo("Success","Entry Deleted Successfully")
        refresh_enquiries()

def turn_on_filter():
    global edit
    for lable in active_data_lables:
        lable.destroy()
    active_data_lables.clear()

    for lable in active_filter_list:
        lable.destroy()
    active_filter_list.clear()

    for lable in button_lables:
        lable.destroy()
    button_lables.clear()

    selected_filters.clear()
    for i in range( len(filters_on) ):
        if filters_on[i].get():
            selected_filters.append(filter_enq_headings[i])

    for col_idx, text_i in enumerate(selected_filters):
        filter_label = CTkLabel(dashboard_frame,text=text_i,font=("Montserrat",15,"bold"),text_color="#333333",justify="center")
        filter_label.grid(row=2,column=col_idx,pady=0,padx=40,sticky="we")
        if text_i == "Sr.No":
            filter_label.grid_configure(sticky="ew",padx=10)
        active_filter_list.append(filter_label)

    a = ["Sr.No","Gender","Name","Course","Source","Mobile No.","Date","Time"]
    try:
        if len(total_enquiries_list) == 0:
            data_lable = CTkLabel(dashboard_frame,text="No Enquies",font=("Montserrat",18),text_color="#000000",justify="center")
            data_lable.grid(row=3,column=0,sticky="we",pady=15,padx=60,columnspan=3)
            edit_button.grid_forget()
            cancel_button.grid_forget()
        else:
            for i , j  in enumerate(total_enquiries_list):
                for col , same in enumerate(selected_filters):
                    if same in a:
                        if same == "Sr.No":
                            data_lable = CTkLabel(dashboard_frame,text=i+1,font=("Montserrat",14),text_color="#000000",justify="center")
                            data_lable.grid(row=3+i,column=col,sticky="we",pady=3)
                            data_lable.configure(justify="center")
                            data_lable.grid_configure(sticky="ew",padx=0)
                        else:
                            data_lable = CTkLabel(dashboard_frame,text=j[a.index(same)],font=("Montserrat",14),text_color="#000000",justify="center")
                            data_lable.grid(row=3+i,column=col,sticky="we",pady=3)
                        active_data_lables.append(data_lable)

                edit_button = CTkButton(dashboard_frame,text="⋮",hover_color="#95FEE0",width=1,font=("Calibri",20),text_color="#333333",fg_color="#cefdfd",corner_radius=90,bg_color="#ecffff",command=lambda temp=i:edit_enquiry_function(temp))
                edit_button.grid(row=3+i,column=len(selected_filters),pady=3)
                button_lables.append(edit_button)
    except:
        pass


default_checked = ["Sr.No","Gender","Name","Course","Source","Mobile No.","Date","Time"]

for i,heading in enumerate(filter_enq_headings):
    var = BooleanVar()
    filters_on.append(var)
    checkbox = CTkCheckBox(filter_enq_headings_frame,text=heading,variable=var,command=turn_on_filter,font=("Calibri",14),text_color="#000000",fg_color="#333333",checkbox_width=18,checkbox_height=18,border_width=1)
    checkbox.grid(row= i // 8,column=i % 8,pady=5,padx=15,sticky="we")
    if heading in default_checked:
        checkbox.select()
        
# --------------------------------------------------------------------------------------------------------------------------------------------------------
wb = load_workbook("Student Enquiries Backup.xlsx")
sheet2 = wb["Student Enquiries"]

total_enquiries_list = []
def refresh_enquiries():
    global wb , sheet2 , total_enquiries_list
    total_enquiries_list.clear()
    wb = load_workbook("Student Enquiries Backup.xlsx")
    sheet2 = wb["Student Enquiries"]
    for row in sheet2.iter_rows(min_row=2,values_only=True):
        sr_no , gender_show,  name , course_show , source_show, edu_show, mobile_show , college_show , address_show , on_date , on_time = row
        total_enquiries_list.append([sr_no,gender_show,name,course_show,source_show,mobile_show,on_date,on_time])
    turn_on_filter()
refresh_enquiries()

# Enquiry Frame
enq_frame = CTkFrame(root,bg_color="#333333",fg_color="#ecffff",width=1000,corner_radius=12,border_color="#21ffe1",border_width=1)
CTkLabel(enq_frame,text="Add Enquiry",font=("Montserrat",16,"bold"),text_color="#ffffff",bg_color="#ecffff",fg_color="#333333",border_color="#5476ff",border_width=1,width=720,anchor="w",corner_radius=12).grid(row=0,column=0,sticky="w",padx=10,pady=10,columnspan=6)
CTkLabel(enq_frame,text="Gender",font=("Montserrat",14,"bold"),text_color="#333333",bg_color="#ecffff",fg_color="#ecffff").grid(row=1,column=0,sticky="w",padx=(20,10),pady=0)
CTkLabel(enq_frame,text="First Name*",font=("Montserrat",14,"bold"),text_color="#333333",bg_color="#ecffff",fg_color="#ecffff").grid(row=1,column=2,sticky="w",padx=(20,10),pady=0)
CTkLabel(enq_frame,text="Last Name*",font=("Montserrat",14,"bold"),text_color="#333333",bg_color="#ecffff",fg_color="#ecffff").grid(row=1,column=4,sticky="w",padx=(20,10),pady=0)
CTkLabel(enq_frame,text="Add Course*",font=("Montserrat",14,"bold"),text_color="#333333",bg_color="#ecffff",fg_color="#ecffff").grid(row=2,column=0,sticky="w",padx=(20,10),pady=0)
CTkLabel(enq_frame,text="Source",font=("Montserrat",14,"bold"),text_color="#333333",bg_color="#ecffff",fg_color="#ecffff").grid(row=3,column=0,sticky="w",padx=(20,10),pady=0)
CTkLabel(enq_frame,text="Qualification",font=("Montserrat",14,"bold"),text_color="#333333",bg_color="#ecffff",fg_color="#ecffff").grid(row=3,column=2,sticky="w",padx=(20,10),pady=0)
CTkLabel(enq_frame,text="Mobile Number*",font=("Montserrat",14,"bold"),text_color="#333333",bg_color="#ecffff",fg_color="#ecffff").grid(row=3,column=4,sticky="w",padx=(20,10),pady=0)
CTkLabel(enq_frame,text="College",font=("Montserrat",14,"bold"),text_color="#333333",bg_color="#ecffff",fg_color="#ecffff").grid(row=4,column=0,sticky="w",padx=(20,10),pady=0)
CTkLabel(enq_frame,text="Address",font=("Montserrat",14,"bold"),text_color="#333333",bg_color="#ecffff",fg_color="#ecffff").grid(row=5,column=0,sticky="w",padx=(20,10),pady=0)

gender_enq = StringVar()
CTkComboBox(enq_frame,font=("Montserrat",12),values=["Male","Female","Other"],text_color="#ffffff",fg_color="#333333",width=150,variable=gender_enq).grid(row=1,column=1)

first_name = CTkEntry(enq_frame,font=("Montserrat",14),text_color="#ffffff",fg_color="#333333",width=150)
first_name.grid(row=1,column=3)

last_name = CTkEntry(enq_frame,font=("Montserrat",14),text_color="#ffffff",fg_color="#333333",width=150)
last_name.grid(row=1,column=5,padx=(0,20))

# Mutiple Courses List in Enquiry Form
enq_course = StringVar()
enq_courses_list = []
clean_str = ""
def courses_list_func(args):
    global clean_str
    if enq_course.get() not in enq_courses_list:
        enq_courses_list.append(enq_course.get())

    clean_str = ", ".join(enq_courses_list)
    courses_label.configure(text=clean_str)

CTkComboBox(enq_frame,font=("Montserrat",12),variable=enq_course,values=["MSCIT","Tally Prime with GST","Adv. Tally Prime","Adv. Excel Pro","Python","C,C++"],command=courses_list_func,text_color="#ffffff",fg_color="#333333",width=150).grid(row=2,column=1,pady=10)
courses_label = CTkLabel(enq_frame,text="",corner_radius=6,anchor="w",font=("Calibri",15),text_color="#000000",bg_color="#ecffff",fg_color="#ecffff",border_color="#333333",border_width=1)
courses_label.grid(row=2,column=2,columnspan=4,sticky="ew",padx=(10,20))

enq_source = StringVar()
CTkComboBox(enq_frame,font=("Montserrat",12),variable=enq_source,values=["Reference","Walk-in","Social Media","Ex-Student","Google Form"],text_color="#ffffff",fg_color="#333333",width=150).grid(row=3,column=1)

enq_education = StringVar()
CTkComboBox(enq_frame,font=("Montserrat",12),variable=enq_education,values=["School Student","Jr. College Stuent","College Student","Other"],text_color="#ffffff",fg_color="#333333",width=150).grid(row=3,column=3)

enq_phone = CTkEntry(enq_frame,font=("Montserrat",14),text_color="#ffffff",fg_color="#333333",width=150)
enq_phone.grid(row=3,column=5,padx=(0,20))

enq_college = CTkEntry(enq_frame,font=("Montserrat",14),text_color="#ffffff",fg_color="#333333",width=150)
enq_college.grid(row=4,column=1,pady=10,columnspan=4,sticky="we")

enq_address = CTkEntry(enq_frame,font=("Montserrat",14),text_color="#ffffff",fg_color="#333333",width=150)
enq_address.grid(row=5,column=1,columnspan=4,sticky="we")

def save_enquiry():
    global clean_str , wb , sheet2
    gender_enq_get = gender_enq.get()
    enq_source_get= enq_source.get()
    enq_education_get = enq_education.get()
    enq_college_get = enq_college.get()
    enq_address_get = enq_address.get()
    full_name = f"{first_name.get()} {last_name.get()}"

    if len(full_name) == 0 or len(clean_str) == 0:
        msg.showwarning("Oops!","Please fill the compulsory information first")
        return

    try:
        enq_phone_get = int(enq_phone.get())
    except ValueError:
        msg.showwarning("Oops!","Please Enter Phone Number")
        return

    wb = load_workbook("Student Enquiries Backup.xlsx")
    sheet2 = wb["Student Enquiries"]
    sr_no_show = sheet2.max_row

    get_date_time() # To Accept Time
    sheet2.append([sr_no_show,gender_enq_get,full_name,clean_str,enq_source_get,enq_education_get,enq_phone_get,enq_college_get,enq_address_get,current_date,current_time])
    try:
        os.system("tskill excel")
    except:
        pass
    wb.save("Student Enquiries Backup.xlsx")

    shutil.copy(os.path.join(home,"KCA-Student Login","Student Enquiries Backup.xlsx"),os.path.join(home,"Documents","Student Enquiries.xlsx"))
    refresh_enquiries()
    msg.showinfo("Enquiry Saved!","Enquiry Saved Successfully!")

    gender_enq.set("")
    first_name.delete(0,"end")
    last_name.delete(0,"end")
    courses_label.configure(text="")
    clean_str=""
    enq_course.set("")
    enq_source.set("")
    enq_education.set("")
    enq_phone.delete(0,"end")
    enq_college.delete(0,"end")
    enq_address.delete(0,"end")
    enq_courses_list.clear()

CTkButton(enq_frame,text="Save Enquiry",font=("Montserrat",14,"bold"),corner_radius=12,command=save_enquiry).grid(row=6,column=0,columnspan=6,sticky="we",padx=10,pady=10)

# -----------------------------------------------------------------------------------------------------------------------------------------------------
#(Discount Calulation Displaying Window >>> )
total = 0
inst_fees = 0
def auto(*args):
    global total
    global course_fee , inst_fees
    type_input = dis_var.get()
    if not type_input:
        discount_amt = 0
    else:
        try:
            discount_amt = int(type_input)
        except:
            return

    total = course_fee - discount_amt
    inst_fees = total
    label3.configure(text=f"₹ {max(0,total)} /-")
# -----------------------------------------------------------------------------------------------------------------------------------------------------

# Submit Button - Saving data Entry
def save_entry():
    global total
    global course_fee
    global label4
    try : 
        first = e1.get().strip()
        middle = e2.get().strip()
        last = e3.get().strip()
        gender = g.get().strip()
        academic_status = std.get().strip()
        class_year = std2.get().strip()
        board = accept_board.get().strip()
        course = accept_course.get().strip()
        day = int(divas.get())
        month = m.get()
        year = y.get()
        college = e4.get().strip()
        tution = e5.get().strip()
    except :
        msg.showwarning("Oops!","Please fill the information first.")
        return
    mobile1 = m1.get()
    mobile2 = m2.get()
    student_email = stu_mail.get().strip()
    address = add.get().strip()


    if len(mobile1) == 0:
        msg.showerror("Must Insert","You must insert contact Number first")
        return
    elif len(mobile1) != 10:
        msg.showwarning("Invalid No","You entered an incorrect phone number")
        return

    wb = load_workbook("Student Enquiries Backup.xlsx")
    s = wb.active
    name = f"{first} {middle} {last}"
    row = s.max_row

    match day:
        case 1 | 21 | 31 :
            dob = f"{day}st {month} {year}"
        case 2 | 22 :
            dob = f"{day}nd {month} {year}"
        case 3 | 23 :
            dob = f"{day}rd {month} {year}"
        case _:
            dob = f"{day}th {month} {year}"

    installment_values = []
    for i in entry_vars:
        val_str = i.get().strip()
        try:
            installment_values.append(int(val_str))
        except ValueError:
            installment_values.append(0)

    if total == 0:
        total = course_fee
 
    wb = load_workbook("Student Enquiries Backup.xlsx")
    s = wb["Admissions"]
    get_date_time()
    if std.get() == "School Student" or std.get() == "Jr. College" :
        if std.get() == "Jr. College":
            class_year = f"{class_year} ({jr_stream.get()})"
        try:
            s.append([row,name,int(mobile1),gender,academic_status,class_year,board,course,total,dob,college,tution,int(mobile2),student_email,address,current_date,current_time] + installment_values)
        except:
            s.append([row,name,int(mobile1),gender,academic_status,class_year,board,course,total,dob,college,tution,"-",student_email,address,current_date,current_time] + installment_values)

    elif std.get() == "Other" :
        class_year = e7.get().strip()
        try:
            s.append([row,name,int(mobile1),gender,academic_status,class_year,"-",course,total,dob,college,tution,int(mobile2),student_email,address,current_date,current_time]+ installment_values)
        except:
            s.append([row,name,int(mobile1),gender,academic_status,class_year,"-",course,total,dob,college,tution,"-",student_email,address,current_date,current_time]+ installment_values)
    else:
        stream = e7.get().strip()
        class_year = f"{class_year} ({stream})"
        try:
            s.append([row,name,int(mobile1),gender,academic_status,class_year,"-",course,total,dob,college,tution,int(mobile2),student_email,address,current_date,current_time]+ installment_values)
        except:
            s.append([row,name,int(mobile1),gender,academic_status,class_year,"-",course,total,dob,college,tution,"-",student_email,address,current_date,current_time]+ installment_values)

    try:
        os.system("tskill excel")
    except:
        pass

    wb.save("Student Enquiries Backup.xlsx")
    shutil.copy(os.path.join(home,"KCA-Student Login","Student Enquiries Backup.xlsx"),os.path.join(home,"Documents"))
    os.replace(os.path.join(home,"Documents","Student Enquiries.xlsx"),os.path.join(home,"Documents","Student Enquiries.xlsx"))
    msg.showinfo("Success!","Entry Sucessfully added.")
        
    e1.delete(0,"end")
    e2.delete(0,"end")
    e3.delete(0,"end")
    g.set("")
    std.set("")
    accept_course.set("")
    divas.set("")
    m.set("")
    y.set("")
    accept_board.set("")
    e4.delete(0,"end")
    e5.delete(0,"end")
    m1.delete(0,"end")
    m2.delete(0,"end")
    stu_mail.delete(0,"end")
    add.delete(0,"end")
    label.configure(text=f"₹ {0} /-",font=("Montserrat",18))
    try :
        label4.grid_remove()
    except:
        pass
    e6.delete(0,"end")
    e7.delete(0,"end")

    #  Gmail Integration
    server = smtplib.SMTP("smtp.gmail.com",587)
    server.starttls()

    with open("login_creadentials.txt","r") as file :
        first_line , second_line = file.read().strip().split("|")
        message = f"""Subject: Admission Confirmation - Kailash Computer Academy\n\n
Hi {first},
Congratulations! We are pleased to inform you that your admission at Kailash Computer Academy for {course} has been successfully confirmed.
We are excited to have you join us and help you achieve your academic goals.

Best regards,
Kailash Computer Academy
(Gov. Authorised Learning Center)"""
    try :
        server.login(first_line,second_line)
        server.sendmail(first_line,student_email,message)
    except:
        gm = msg.askretrycancel("Oops!","Your email isn't Valid!\nPlease enter valid email address")

# -----------------------------------------------------------------------------------------------------------------------------------------------------

# Student Information Frame
f1 = CTkScrollableFrame(root,width=1060,height=180,label_font=("Montserrat",18),label_text_color="#ffffff",label_fg_color="#333333",bg_color="#333333",corner_radius=12,label_text="Student Information",fg_color="#ECffff",border_color="#21FFE1",border_width=1)
# f1.grid(row=1, column=0, sticky="ew", padx=10, pady=(5,5))

# Vertical Side Frame
ver_frame = CTkFrame(root,width=266,bg_color="#333333",fg_color="#ECFFFF",corner_radius=12,border_color="#21FFE1",border_width=1)
ver_frame.grid(row=0,column=1,rowspan=4,sticky="wens",pady=10,padx=(0,20))

# KCA LOGO IMAGE
# logo = CTkImage( light_image= Image.open(r"D:\Sagar - KCA 4\Python Course\Python Programs\logo.jpg"),
#                 dark_image= Image.open(r"D:\Sagar - KCA 4\Python Course\Python Programs\logo.jpg"),
#                 size=(150,70))

# Names
CTkLabel(f1,text="First Name*",font=("Montserrat",15),text_color="#000000",corner_radius=5,padx=9,justify="left").grid(sticky="w",row=0,column=0,padx=(0,3),pady=0)
e1 = CTkEntry(f1,border_width=1,width=180,height=27,font=("Montserrat",14),fg_color="#333333",corner_radius=6,text_color="#FFFFFF")
e1.grid(row=0,column=1,sticky="w")

CTkLabel(f1,text="Middle Name",font=("Montserrat",15),text_color="#000000",corner_radius=5,padx=9).grid(sticky="w",row=0,column=2,padx=3,pady=0)
e2 = CTkEntry(f1,border_width=1,width=180,height=27,font=("Montserrat",14),fg_color="#333333",corner_radius=6,text_color="#FFFFFF")
e2.grid(row=0,column=3,sticky="w")

CTkLabel(f1,text="Last Name*",font=("Montserrat",15),height=27,text_color="#000000",corner_radius=5,padx=0).grid(sticky="w",row=0,column=4,padx=10,pady=0)
e3 = CTkEntry(f1,border_width=1,width=180,font=("Montserrat",14),fg_color="#333333",corner_radius=6,text_color="#FFFFFF")
e3.grid(row=0,column=5,sticky="w")

# Gender
g = StringVar()
CTkLabel(f1,text="Gender*",font=("Montserrat",15),text_color="#000000",corner_radius=5,padx=9,justify="left").grid(sticky="w",row=1,column=0,padx=0,pady=5)
CTkComboBox(f1,variable=g,values=["Male","Female","Other"],width=180,fg_color="#333333",text_color="#FFFFFF" ).grid(row=1,column=1,sticky="w",pady=0)

std2 = StringVar()          # Line No. 344 , 355 , 356 It is Declared
stm = StringVar()          # Line No. 344 , 355 , 356 It is Declared
def std_trigger(*args):
    global e7 , s2
    school = [ "5th" , "6th" , "7th" , "8th" , "9th" , "10th" ]
    jr_college = [ "11th" , "12th" ]
    college = ["FY - TY","Completed"] 
    b_lable.grid(sticky="w",row=0,column=4,padx=(220,20),pady=(5,0))
    b.grid(row=0,column=5,sticky="w")

    e7.grid_forget()
    s2.grid_forget()
    jr_stream.grid_forget()

    if std.get() == "School Student":
        s2.configure(values=school)
        s2.grid(row=0,column=4,sticky="w",padx=(35,0))
        
    elif std.get() == "Jr. College":
        s2.configure(values=jr_college,width=70)
        s2.grid(row=0,column=4,sticky="w",padx=(35,0))
        jr_stream.configure(values=["Science","Commerce","Arts"],width=102)
        jr_stream.grid(row=0,column=4,sticky="w",padx=(114,0))

    elif std.get() == "College Student":
        s2.configure(values=college,width=100)
        s2.grid(row=0,column=4,sticky="w",padx=(35,0),pady=(2,3))
        e7.configure(width=79)
        e7.grid(row=0,column=4,sticky="w",padx=(139,10),pady=(2,3))
        b.grid_remove()
        b_lable.grid_remove()
    else:
        e7.configure(width=140)
        e7.grid(row=0,column=4,sticky="w",padx=(35,0),pady=(2,3))
        b.grid_remove()
        b_lable.grid_remove()

# Academic Status
std = StringVar()
CTkLabel(f1,text="Academic Status*",font=("Montserrat",15),text_color="#000000",corner_radius=5,padx=9,justify="left").grid(sticky="w",row=1,column=2,padx=3,pady=8)
CTkComboBox(f1,variable=std,values=["School Student","Jr. College","College Student","Other"],width=180,fg_color="#333333",text_color="#FFFFFF",command=std_trigger).grid(row=1,column=3,sticky="w")

# Course Fees Calculation
accept_course = StringVar()
accept_board = StringVar()
label = CTkLabel(f1,width=110,text=f"₹ {0}",font=("Montserrat",15),text_color="#ffffff",corner_radius=10,fg_color="#333333",border_color="#00D5FF",border_width=1,pady=2,padx=10)
label.grid(sticky="w",row=3,column=5,padx=(0,3),pady=(5,0))
course_fee = 0

def trigger_fees(value):
    global course_fee , inst_fees
    if accept_course.get() == "12th CS" and accept_board.get() == "STATE":
        course_fee = 10000
    elif accept_course.get() == "12th CS" and accept_board.get() == "CBSE":
        course_fee = 15000
    elif accept_course.get() == "12th English Free Classes" and accept_board.get() == "STATE":
        course_fee = 5500
    elif accept_course.get() == "12th English Free Classes" and accept_board.get() == "CBSE":
        course_fee = 6500
    elif accept_course.get() == "12th IT":
        course_fee = 5000
    else:
        for i in range(len(course_names)):
            if course_names[i] == accept_course.get():
                course_fee = int(course_fees[i])

    label.configure(text=f"₹ {course_fee} /-")
    label2.configure(text=f"₹ {course_fee} /-")
    final_fees.configure(text=f"₹ {course_fee} /-")
    inst_fees = course_fee

# To Add Courses Text File
if not os.path.exists("Courses.txt"):
    with open("Courses.txt","a") as file:
        pass

course_names = []
course_fees = []
# Course
cr = StringVar()
CTkLabel(f1,text="Course*",font=("Montserrat",15),text_color="#000000",corner_radius=5).grid(sticky="w",row=1,column=4,padx=(10,3),pady=8)
combo1 = CTkComboBox(f1,variable=accept_course,width=180,fg_color="#333333",text_color="#FFFFFF",command=trigger_fees)
combo1.grid(row=1,column=5,sticky="w")

def refresh_courses():
    course_names.clear()
    course_fees.clear()
    global multi_course
    multi_course = ["12th CS","12th English Free Classes","12th IT"]
    with open("Courses.txt","r") as file:
        lines = file.readlines()
        for i in lines:
            if "|" in i:
                course_names.append(i[ : i.index("|")])
                course_fees.append(i[i.index("|")+1 : len(i)-1])        

    multi_course.extend(course_names)
    combo1.configure(values=multi_course)
refresh_courses()

def add_course():
    global r
    try:
        if r.winfo_exists():
            r.destroy()
    except:
        pass 
    
    r = CTk()
    r.title("Course Alteration")
    r.config(bg="#FFFFFF")

    def add_frame_function():
        delete_frame.grid_forget()
        add_frame.grid(row=1,column=0,columnspan=2,padx=10,pady=10,sticky="ew")
        b1.configure(fg_color="#a0f5fc")
        b2.configure(fg_color="#00eaff")

    def delete_frame_function():
        add_frame.grid_forget()
        delete_frame.grid(row=1,column=0,columnspan=2,padx=10,pady=10,sticky="ew")
        b1.configure(fg_color="#00eaff")
        b2.configure(fg_color="#a0f5fc")

    b1 = CTkButton(r,font=("Montserrat",15),text="Add Course",corner_radius=5,bg_color="#ffffff",fg_color="#77a4ff",text_color="#000000",width=200,command=add_frame_function)
    b1.grid(row=0,column=0,pady=(10,5),padx=10,sticky="w")
    b2 = CTkButton(r,font=("Montserrat",15),text="Delete Course",corner_radius=5,bg_color="#ffffff",fg_color="#77a4ff",text_color="#000000",width=200,command=delete_frame_function)
    b2.grid(row=0,column=1,pady=(10,5),padx=10,sticky="w")

    add_frame = CTkFrame(r,bg_color="#ffffff",fg_color="#ffffff",border_width=1)
    delete_frame = CTkFrame(r,bg_color="#ffffff",fg_color="#ffffff",border_width=1)

    # Add Frame Contents
    CTkLabel(add_frame,text="Course Name",font=("Montserrat",15),bg_color="#ffffff",text_color="#000000").grid(row=0,column=0,padx=10,sticky="w",pady=0)
    CTkLabel(add_frame,text="Course Fees",font=("Montserrat",15),bg_color="#ffffff",text_color="#000000").grid(row=1,column=0,padx=10,sticky="w",pady=(0,5))

    c1=CTkEntry(add_frame,font=("Montserrat",15),bg_color="#FFFFFF",text_color="#000000",fg_color="#e0e0e0",border_width=1,border_color="#3acaff",width=250)
    c1.grid(row=0,column=1,padx=20,pady=10)
    c2=CTkEntry(add_frame,font=("Montserrat",15),bg_color="#FFFFFF",text_color="#000000",fg_color="#e0e0e0",border_width=1,border_color="#3acaff",width=250)
    c2.grid(row=1,column=1,padx=20,pady=(0,10))

    # Delete Frame Contents
    if len(course_names) == 0:
        CTkLabel(delete_frame,text="No Course Added ",font=("Montserrat",15),bg_color="transparent",text_color="#000000").grid(row=1,column=0,columnspan=3,padx=(30,35),sticky="w",pady=0)
        
    CTkLabel(delete_frame,text="Select",font=("Montserrat",15,"bold"),bg_color="transparent",text_color="#000000").grid(row=0,column=0,padx=(30,35),sticky="w",pady=0)
    CTkLabel(delete_frame,text="Course Names",font=("Montserrat",15,"bold"),bg_color="transparent",text_color="#000000").grid(row=0,column=1,padx=(0,45),sticky="w",pady=5)
    CTkLabel(delete_frame,text="Course Fees",font=("Montserrat",15,"bold"),bg_color="transparent",text_color="#000000").grid(row=0,column=2,padx=(10,0),sticky="w",pady=0)

    delete_checkbox = []
    for i in range(len(multi_course)-3):
        row_tracker = BooleanVar(value=False)
        delete_checkbox.append(row_tracker)

        CTkCheckBox(delete_frame,text=" ",border_width=1,width=1.5,height=1.5,variable=row_tracker).grid(row=i+1,column=0)
        CTkLabel(delete_frame,text=course_names[i],font=("Montserrat",12),bg_color="#ffffff",text_color="#000000").grid(row=i+1,column=1,padx=(10,0),sticky="w",pady=0)
        CTkLabel(delete_frame,text=f"₹ {course_fees[i]} /-",justify="center",font=("Montserrat",14),bg_color="#ffffff",text_color="#000000").grid(row=i+1,column=2,padx=10,sticky="w",pady=0)

    def confirm_deletion():
        remaining_lines = []
        for i in range(len(delete_checkbox)): 
            if not delete_checkbox[i].get():
                remaining_lines.append(f"{course_names[i]}|{course_fees[i]}\n")

        with open("Courses.txt","w") as file:
            file.writelines(remaining_lines)
        refresh_courses()
        msg.showinfo("Success!","Course Deletation Successfull!")
    CTkButton(delete_frame,font=("Montserrat",15),text="Confirm",width=200,command=confirm_deletion).grid(row=50,column=0,columnspan=3,pady=10,padx=10,sticky="we")

    def added():
        course = c1.get()
        course_fee = int(c2.get())
        with open("Courses.txt","r") as file:
            lines = file.readlines()
        this_course = f"{course}|{course_fee}\n"

        if this_course in lines:
            msg.showwarning("Course Exists!","This course already exits!")
        else:
            with open("Courses.txt","a") as file:
                file.write(this_course)
            msg.showinfo("Success!","Course Added Successfully!")
            r.destroy()
        refresh_courses()
    CTkButton(add_frame,font=("Montserrat",15),text="Add the Course",command=added).grid(row=2,column=0,columnspan=2,pady=10,padx=10,sticky="we")
    r.mainloop()


# Date of Birth
dob_frame= CTkFrame(f1,fg_color="transparent")
dob_frame.grid(row=2,column=0,columnspan=6,padx=0,sticky="w",pady=0)
dob_frame.rowconfigure(0,weight=0)
months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

divas = StringVar()
m = StringVar()
y = StringVar()
CTkLabel(dob_frame,text="Date of Birth*",font=("Montserrat",15),text_color="#000000",corner_radius=5,padx=9).grid(sticky="w",row=0,column=0,padx=(0,10),pady=0)
CTkComboBox(dob_frame,variable=divas,font=("Montserrat",14),text_color="#FFFFFF",fg_color="#333333",values=[str(i) for i in range(1,32)],width=70,height=28).grid(row=0,column=1,sticky="w",padx=(19,0))
CTkComboBox(dob_frame,variable=m,font=("Montserrat",14),text_color="#FFFFFF",fg_color="#333333",values=months,width=120,height=28).grid(row=0,column=2,sticky="w",padx=10)
CTkComboBox(dob_frame,variable=y,font=("Montserrat",14),text_color="#FFFFFF",fg_color="#333333",values=[str(i) for i in range(1990,2027)],width=85,height=28).grid(row=0,column=3,sticky="w")

e7 = CTkEntry(dob_frame,text_color="#ffffff",bg_color="transparent",fg_color="#333333",border_color="#A4F7FF",border_width=1,width=78)
s2 = CTkComboBox(dob_frame,variable=std2,font=("Montserrat",12),text_color="#FFFFFF",fg_color="#333333",width=120,height=28)
jr_stream = CTkComboBox(dob_frame,variable=stm,font=("Montserrat",12),text_color="#FFFFFF",fg_color="#333333",width=120,height=28)

# board
bd = StringVar()
b_lable = CTkLabel(dob_frame,text="Board",font=("Montserrat",15),text_color="#000000",corner_radius=5,justify="left")
b_lable.grid(sticky="w",row=0,column=4,padx=(220,20),pady=(5,0))
b = CTkComboBox(dob_frame,variable=accept_board,values=["STATE","CBSE"],width=180,fg_color="#333333",text_color="#FFFFFF",command=trigger_fees)
b.grid(row=0,column=5,sticky="w",padx=(48,0))

# College Name
CTkLabel(f1,text="College Name",font=("Montserrat",15),text_color="#000000",corner_radius=6,padx=9).grid(sticky="w",row=3,column=0,padx=(0,10),pady=5)
e4 = CTkEntry(f1,border_width=1,width=498,font=("Montserrat",14),fg_color="#333333",corner_radius=6,text_color="#FFFFFF" )
e4.grid(sticky="w",row=3,column=1,padx=(0,10),pady=5,columnspan=4)

# Fees 
CTkLabel(f1,text="Fees",font=("Montserrat",15),text_color="#000000",corner_radius=5).grid(sticky="w",row=3,column=4,padx=5,pady=5)

# Coaching Classes Name
CTkLabel(f1,text="Coaching Class",font=("Montserrat",15),text_color="#000000",corner_radius=6,padx=9).grid(sticky="w",row=4,column=0,padx=(0,10),pady=5)
e5 = CTkEntry(f1,border_width=1,width=498,font=("Montserrat",14),fg_color="#333333",corner_radius=6,text_color="#FFFFFF" )
e5.grid(sticky="w",row=4,column=1,padx=(0,10),pady=5,columnspan=4)

#---------------------------------------------------------------------------------------------------------------------------------------------------------------

# Contact Information
f2 = CTkScrollableFrame(root,width=1050,height=0,label_font=("Montserrat",15,"bold"),label_text_color="#ffffff",corner_radius=12,label_fg_color="#333333",label_text="Contact Information",fg_color="#ECFFFF",bg_color="#333333",border_color="#21FFE1",border_width=1)
# f2.grid(row=2, column=0, sticky="new", padx=10, pady=0)

CTkLabel(f2,text="Mobile 1*",font=("Montserrat",15),text_color="#000000",corner_radius=5,padx=9,justify="left").grid(sticky="w",row=0,column=0,padx=(0,3),pady=5)
m1 = CTkEntry(f2,border_width=1,width=180,font=("Montserrat",14),corner_radius=6,fg_color="#333333",text_color="#FFFFFF")
m1.grid(row=0,column=1,sticky="w",padx=(0,40))

CTkLabel(f2,text="Parent's No.",font=("Montserrat",15),text_color="#000000",corner_radius=5,padx=9,justify="left").grid(sticky="w",row=0,column=2,padx=(0,3),pady=5)
m2 = CTkEntry(f2,border_width=1,width=180,font=("Montserrat",14),corner_radius=6,fg_color="#333333",text_color="#FFFFFF")
m2.grid(row=0,column=3,sticky="w",padx=(0,0))

# Email
CTkLabel(f2,text="Email",font=("Montserrat",15),text_color="#000000",corner_radius=5,padx=9,justify="left").grid(sticky="w",row=1,column=0,padx=(0,3),pady=5)
stu_mail = CTkEntry(f2,border_width=1,width=530,font=("Montserrat",14),corner_radius=6,fg_color="#333333",text_color="#FFFFFF")
stu_mail.grid(row=1,column=1,sticky="w",padx=(0,20),columnspan=4)  

# Address
CTkLabel(f2,text="Address",font=("Montserrat",15),text_color="#000000",corner_radius=5,padx=9,justify="left").grid(sticky="w",row=2,column=0,padx=(0,3),pady=0)
add = CTkEntry(f2,border_width=1,width=530,height=50,font=("Montserrat",14),corner_radius=6,fg_color="#333333",text_color="#FFFFFF")
add.grid(row=2,column=1,sticky="w",padx=(0,0),pady=(8,0),columnspan=3)

CTkButton(f2,text="Submit",corner_radius=150,font=("Montserrat",25),command=save_entry,width=200).grid(row=3,column=0,columnspan=7,sticky="w",padx=0,pady=(10,0))
# img_label1 = CTkLabel(f2,image=logo,text="").grid(row=0,column=5,rowspan=4,pady=(80,0),padx=(230,0))

#---------------------------------------------------------------------------------------------------------------------------------------------------------------
# Discount Code >>>
# Magane Discount Window Show/Hide
def discount_check():
    if check.get():
        dis_frame.grid(row=0,column=0,padx=5,pady=(5,0))
    else:
        dis_frame.grid_remove()

dis_var = StringVar()
dis_var.trace_add("write",auto)

my_strikethrough_font = CTkFont(family="Montserrat", size=18, slant="italic", overstrike=True)
label4 = None

def proceed():
    global label4 , final_fees
    label.configure(text=f"₹ {course_fee} /-",font=my_strikethrough_font)
    label4 = CTkLabel(f1,text=f"₹ {max(0,total)} /-",font=("Montserrat",20,"bold"),text_color="#000000",corner_radius=5)
    label4.grid(sticky="w",row=3,column=5,padx=(120,3),pady=(5,0))
    final_fees.configure(text=f"₹ {total} /-")
    trigger_no_install()

# Discount Fees 
dis_frame = CTkScrollableFrame(ver_frame,width=220,label_text="Manage Discount",label_fg_color="#333333",label_text_color="#ffffff",label_font=("Montserrat",15,"italic"),bg_color="transparent",fg_color="#ecffff",corner_radius=15,border_color="#FF3482",border_width=1.5)
CTkLabel(dis_frame,text="Fees",font=("Montserrat",14,"italic"),text_color="#000000").grid(sticky="w",row=0,column=0,padx=(0,10))
label2 = CTkLabel(dis_frame,width=110,text=course_fee,font=("Montserrat",14),text_color="#FFFFFF",corner_radius=10,fg_color="#333333",border_color="#00D5FF",border_width=1,padx=10)
label2.grid(sticky="w",row=0,column=1,padx=5)

CTkLabel(dis_frame,text="Discount Amt",font=("Montserrat",14,"italic"),text_color="#000000",padx=0).grid(sticky="w",row=1,column=0,padx=(0,10),pady=(5,0))
e6 = CTkEntry(dis_frame,width=110,textvariable=dis_var,text_color="#ffffff",bg_color="transparent",corner_radius=10,fg_color="#333333",border_width=1,border_color="#00D5FF")
e6.grid(row=1,column=1,pady=(5,0),padx=5)

CTkLabel(dis_frame,text="Final Amt",font=("Montserrat",14,"italic"),text_color="#000000",padx=0).grid(sticky="w",row=2,column=0,padx=(0,10),pady=(10,0))
label3 = CTkLabel(dis_frame,width=110,text=course_fee,font=("Montserrat",14),text_color="#FFFFFF",corner_radius=10,fg_color="#333333",border_color="#FF0037",border_width=1,padx=10)
label3.grid(sticky="w",row=2,column=1,padx=5,pady=10)

CTkButton(dis_frame,width=210,text="Proceed",font=("Montserrat",14,"bold"),text_color="#ffffff",corner_radius=12,command=proceed).grid(row=3,column=0,columnspan=2,pady=(10,0))
#------------------------------------------------------------------------------------------------------------------------------

# Installment Window
def installment_window():
    if ins.get():
        ins_frame.grid(row=3,column=0,padx=5,pady=(5,0))

    else:
        ins_frame.grid_remove()

# CheckBox
check = BooleanVar()
ins = BooleanVar()
no_ins = IntVar()
CTkCheckBox(f1,text="Disc. Offer?",font=("Montserrat",15,"italic"),text_color="#3B3B3B",fg_color="#178080",border_width=1,variable=check,checkbox_height=18,checkbox_width=18,command=discount_check).grid(sticky="w",row=4,column=4,padx=(15,3),pady=0)
CTkCheckBox(f1,text="Installments",font=("Montserrat",15,"italic"),text_color="#383838",fg_color="#178080",border_width=1,variable=ins,checkbox_height=18,checkbox_width=18,command=installment_window).grid(sticky="w",row=4,column=5,padx=(15,3),pady=0)

ins_frame = CTkScrollableFrame(ver_frame,width=220,height=80,label_text="Manage Installements",label_fg_color="#333333",label_text_color="#ffffff",label_font=("Montserrat",15,"italic"),bg_color="transparent",fg_color="#ecffff",corner_radius=15,border_color="#FF3482",border_width=1.5)

CTkLabel(ins_frame,text="Final Fees",font=("Montserrat",14,"italic"),fg_color="transparent",text_color="#050505",corner_radius=15).grid(row=1,column=0,sticky="w",padx=0)
final_fees = CTkLabel(ins_frame,text=f"₹ {total} /-",width=100,font=("Montserrat",14,"italic"),bg_color="transparent",border_width=1,border_color="#FF3E58",fg_color="#333333",text_color="#ffffff",corner_radius=15)
final_fees.grid(row=1,column=1,sticky="w",padx=(0,0),pady=(0,5))

install_label = CTkLabel(ins_frame,text="Installments",font=("Montserrat",14,"italic"),fg_color="transparent",text_color="#050505",corner_radius=15)

entry_vars = []  # Reset the list for the new rounds
is_updating = False 
def trigger_no_install(*args):
    global entry_vars
    for widget in ins_frame.grid_slaves():
        if int(widget.grid_info().get("row", 0)) > 2:
            widget.destroy()

    num = no_ins.get()
    entry_vars = []
    try :
        per_ins = int(inst_fees) // num
        remainder = int(inst_fees) % num
    except:
        pass

    for i in range(1,num+1):
        suffix = "st" if i==1 else "nd" if i==2 else "rd" if i==3 else "th"
        text = f"{i}{suffix}"
        CTkLabel(ins_frame,font=("Montserrat",14,"italic"),text=text,fg_color="transparent",text_color="#050505",corner_radius=15,justify="left").grid(row=i+2,column=0,sticky="w",padx=0,pady=0)
        var = StringVar()
        initial_val = per_ins + (remainder if i == 1 else 0) # Put math remainder in 1st box
        var.set(str(initial_val))
        
        var.trace_add("write", lambda *a, idx=i-1: balance_all_installments(idx))
        entry_vars.append(var)
        e8 = CTkEntry(ins_frame,font=("Montserrat",14,"italic"),textvariable=var,fg_color="#333333",bg_color="transparent",text_color="#FFFFFF",border_width=1,border_color="#2BFFD1",width=100,corner_radius=15,justify="center")
        e8.grid(row=i+2,column=1,sticky="w",padx=0,pady=(0,5))
        
def balance_all_installments(modified_index):
    global is_updating, entry_vars
    if is_updating or len(entry_vars) <= 1: 
        return

    try:
        total_fees = int(inst_fees)
        previous_boxes_total = 0
        for i in range(modified_index):
            val = entry_vars[i].get()
            previous_boxes_total += int(val) if val.isdigit() else 0
            
        typed_str = entry_vars[modified_index].get()
        typed_val = int(typed_str) if typed_str.isdigit() else 0
        remaining_fees = total_fees - previous_boxes_total - typed_val
        
        other_indices = [i for i in range(modified_index + 1, len(entry_vars))]
        num_others = len(other_indices)
        
        if num_others == 0:
            target_idx = modified_index - 1
            current_target_val = int(entry_vars[target_idx].get()) if entry_vars[target_idx].get().isdigit() else 0
            
            assigned = sum(int(v.get()) if v.get().isdigit() else 0 for v in entry_vars)
            diff = total_fees - assigned
            
            is_updating = True
            entry_vars[target_idx].set(str(max(0, current_target_val + diff)))
            is_updating = False
            return

        split_val = max(0, remaining_fees // num_others)
        remainder = max(0, remaining_fees % num_others)
        is_updating = True
        
        for order, idx in enumerate(other_indices):
            # Pass any rounding remainder to the first available upcoming box
            current_box_share = split_val + (remainder if order == 0 else 0)
            entry_vars[idx].set(str(current_box_share))
        is_updating = False 
    except Exception:
        is_updating = False

install_label.grid(row=2,column=0,sticky="w",padx=0,pady=(0,10))
CTkComboBox(ins_frame,font=("Montserrat",14,"italic"),variable=no_ins,values=[str(i)for i in range(1,9)],width=100,bg_color="transparent",fg_color="#333333",text_color="#ffffff",border_color="#FF3E58",border_width=1,corner_radius=15,justify="center",command=trigger_no_install).grid(row=2,column=1,sticky="w",padx=0,pady=(0,10))

#--------------------------------------------------------------------------------------------------------------------------------------------------
# Bottom Frame
bottom_frame = CTkScrollableFrame(root,label_text="",bg_color="#333333",fg_color="#AFFFFF",width=400,height=50,corner_radius=12,border_color="#FFF458",border_width=1)
bottom_frame.grid(row=3,column=0,sticky="ew",columnspan=1,padx=10,pady=5)

def open_excel():
    os.startfile(os.path.join(home,"Documents","Student Enquiries.xlsx"))

def again_login():
    if msg.askyesno("Login Creadentials?","Do You want to change Login Credentials?"):
        login_form_window()

CTkButton(bottom_frame,text="Fetch Excel File",width=160,font=("Montserrat",14,"italic"),corner_radius=15,border_color="#4AFFAB",border_width=1,fg_color="#333333",command=open_excel).grid(row=0,column=0,padx=20)
CTkButton(bottom_frame,text="Change Login Email ",width=160,font=("Montserrat",14,"italic"),corner_radius=15,border_color="#4AFFAB",border_width=1,fg_color="#333333",command=again_login).grid(row=0,column=1,pady=0)
CTkButton(bottom_frame,text="Courses ",width=120,font=("Montserrat",14,"italic"),corner_radius=15,border_color="#4AFFAB",border_width=1,fg_color="#333333",command=add_course).grid(row=0,column=2,pady=0,padx=(20,0))
CTkButton(bottom_frame,text="Refresh ",width=120,font=("Montserrat",14,"italic"),corner_radius=15,border_color="#4AFFAB",border_width=1,fg_color="#333333").grid(row=0,column=3,pady=0,padx=(20,0))

# img_label1 = CTkLabel(bottom_frame,image=logo,text="").grid(row=0,column=5,rowspan=4,pady=(0,0),padx=(20,0),sticky="s")
root.mainloop()